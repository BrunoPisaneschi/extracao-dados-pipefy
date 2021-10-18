from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from os import path, makedirs

from unidecode import unidecode


class excel():
    def __init__(self, file_path=".\\database\\",
                 file_name=datetime.now().strftime("%d_%m_%Y") + "_metricas", file_extension=".xlsx"):
        '''When you create Instance of excel, you can change file_path, file_name and file_extension'''
        self._wb = None
        self._ws = None
        self._file_path = file_path
        self._file_name = file_name
        self._file_extension = file_extension
        self._dict_datas = {}

    def _configure_header(self):
        font = Font(color="FFFFFF", bold=True, size=12)
        fill = PatternFill("solid", fgColor="357ae8")
        alignment = Alignment(horizontal="center", vertical="center")
        self._ws.cell(row=1, column=1).font = font
        self._ws.cell(row=1, column=2).font = font
        self._ws.cell(row=1, column=3).font = font
        self._ws.cell(row=1, column=4).font = font
        self._ws.cell(row=1, column=5).font = font
        self._ws.cell(row=1, column=1).fill = fill
        self._ws.cell(row=1, column=2).fill = fill
        self._ws.cell(row=1, column=3).fill = fill
        self._ws.cell(row=1, column=4).fill = fill
        self._ws.cell(row=1, column=5).fill = fill
        self._ws.cell(row=1, column=1).alignment = alignment
        self._ws.cell(row=1, column=2).alignment = alignment
        self._ws.cell(row=1, column=3).alignment = alignment
        self._ws.cell(row=1, column=4).alignment = alignment
        self._ws.cell(row=1, column=5).alignment = alignment

    def read_excel(self, db):
        self._wb = load_workbook(self._file_path + self._file_name + self._file_extension)
        self._ws = self._wb.active
        _rows = self._ws.rows
        _count = 1
        for _row in _rows:
            if ('titulo' in unidecode(_row[2].value.lower())):
                continue
            if (db == True):
                self._dict_datas[_count] = {'id': _row[0].value, 'id_card': _row[1].value, 'titulo': _row[2].value,
                                            'vaga': _row[3].value, 'nivel': _row[4].value, 'motivo_recusa': '',
                                            'soube_vaga': _row[5].value}
                _count += 1

            else:
                if (_row[3].value is None):
                    self._dict_datas[_count] = {'titulo': _row[2].value, 'vaga': _row[3].value, 'nivel': _row[4].value,
                                                'motivo_recusa': '', 'soube_vaga': _row[5].value}
                    _count += 1
                else:
                    for _recusa in _row[3].value.split(","):
                        self._dict_datas[_count] = {'titulo': _row[2].value, 'vaga': _row[3].value,
                                                    'nivel': _row[4].value,
                                                    'motivo_recusa': _recusa.strip(), 'soube_vaga': _row[5].value}
                        _count += 1
        self._wb.close()
        return self._dict_datas

    def get_last_row(self):
        self._dict_datas = dict_datas
        self._wb = load_workbook(self._file_path + self._file_name + self._file_extension)
        self._ws = self._wb.active
        _last_row = self._ws.max_row
        return self._ws[_last_row]

    def write_excel(self):
        '''Write Excel receive a dictionary'''
        self._wb = Workbook()
        self._ws = self._wb.active
        self._ws.cell(row=1, column=1).value = 'Título'
        self._ws.cell(row=1, column=2).value = 'Vaga'
        self._ws.cell(row=1, column=3).value = 'Nível'
        self._ws.cell(row=1, column=4).value = 'Motivo Recusa'
        self._ws.cell(row=1, column=5).value = 'Como soube da vaga?'
        self._configure_header()

        for index, data in self._dict_datas.items():
            self._ws.cell(row=int(index) + 1, column=1).value = data['titulo']
            self._ws.cell(row=int(index) + 1, column=2).value = data['vaga']
            self._ws.cell(row=int(index) + 1, column=3).value = data['nivel']
            self._ws.cell(row=int(index) + 1, column=4).value = data['motivo_recusa']
            self._ws.cell(row=int(index) + 1, column=5).value = data['soube_vaga']

        # Save the file
        if (not path.exists(self._file_path)):
            makedirs(self._file_path)

        self._ws.title = 'Dashboard'
        self._wb.save(self._file_path + self._file_name + "_2" + self._file_extension)
        return self._file_path + self._file_name + self._file_extension

    def update_excel(self, dict_datas):
        self._dict_datas = dict_datas
        self._wb = load_workbook(self._file_path + self._file_name + self._file_extension)
        self._ws = self._wb.active
        _max_row = self._ws.max_row

        if (len(dict_datas.keys()) > 5):
            self._ws.title = 'Report'
            self._ws.cell(row=int(_max_row), column=1).value = _max_row
            self._ws.cell(row=int(_max_row), column=2).value = self._dict_datas['id_card']
            self._ws.cell(row=int(_max_row), column=3).value = self._dict_datas['titulo']
            self._ws.cell(row=int(_max_row), column=4).value = self._dict_datas['vaga']
            self._ws.cell(row=int(_max_row), column=5).value = self._dict_datas['nivel']
            self._ws.cell(row=int(_max_row), column=6).value = self._dict_datas['motivo_recusa']
            self._ws.cell(row=int(_max_row), column=7).value = self._dict_datas['soube_vaga']
        else:
            self._ws.title = 'Dashboard'
            for index, data in self._dict_datas.items():
                self._ws.cell(row=int(index) + _max_row, column=1).value = data['titulo']
                self._ws.cell(row=int(index) + _max_row, column=2).value = data['vaga']
                self._ws.cell(row=int(index) + _max_row, column=3).value = data['nivel']
                self._ws.cell(row=int(index) + _max_row, column=4).value = data['motivo_recusa']
                self._ws.cell(row=int(index) + _max_row, column=5).value = data['soube_vaga']

        # Save the file
        if (not path.exists(self._file_path)):
            makedirs(self._file_path)

        self._wb.save(self._file_path + self._file_name + self._file_extension)
        self._wb.close()
        return self._file_path + self._file_name + self._file_extension
